

# Python program to read an excel file
""""""
# import openpyxl module
import openpyxl

# Give the location of the file
path = "C:\\Users\\SUMANTH\\Desktop\\demo.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
for row in range(1,sheet_obj.max_row+1):
    for column in range(1, sheet_obj.max_column+1):
        cell_obj = sheet_obj.cell(row = row, column = column)
        print(cell_obj.value, end= " ")
    print("\n")
